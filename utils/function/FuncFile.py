# Library
import json
import logging
from pathlib import Path
from openpyxl import load_workbook
import csv
from typing import List, Union, Dict, Any
from pydantic import BaseModel


# Utils
from utils.constant import FOLDER_INPUT, FOLDER_OUTPUT, FOLDER_TEMP, GITHUB_MAX_PER_PAGE
from utils.model.Igithub_issues import IGitHubIssue
from utils.model.Iguthub_comments import IGitHubComment
from utils.model.Sgithub import TGitHubIssueCommentExport, TGitHubIssueExport


def extract_urls_from_excel(filename: str):
    """
    Extracts all URLs from an Excel (.xlsx) file

    Args:
        file_path (str): Path to the Excel file

    Returns:
        list: Unique list of URLs found in the workbook
    """

    # Create output directory if it doesn't exist

    input_file = f"{FOLDER_INPUT}/{filename}"
    print(f'📥 READ_XLSX "{input_file}"')

    wb = load_workbook(filename=input_file)
    urls = set()

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                # Check for explicit hyperlinks
                if cell.hyperlink:
                    urls.add(cell.hyperlink.target)

                # Check for text that looks like a URL
                if isinstance(cell.value, str) and cell.value.startswith(
                    ("http://", "https://")
                ):
                    urls.add(cell.value.strip())

    result = list(urls)
    print(f'🔗 Found "{len(urls)}" URLs')
    # for index, url in enumerate(result):
    #     print(f"{index}. {url}")

    return result


def import_json(path: str) -> any:
    """
    Import data from a JSON file.

    Args:
        file_path: The path to the JSON file.

    Return the parsed JSON data.
    """
    file_path = Path(path)

    result: any = None
    if not file_path.exists():
        logging.info(f'JSON_NOT_EXISTING import_json("{file_path}")')
        return result

    with open(file_path, "r", encoding="utf-8") as f:
        try:
            result = json.load(f)
        except json.JSONDecodeError as e:
            logging.info(f'JSON_INVALID import_json("{file_path}") : {e.doc, e.pos}')

    # print(f'📥 JSON_IMPORTED "{file_path}"')
    return result


def import_json_issues(name: str, page: int) -> list[IGitHubIssue] | None:
    result: list[IGitHubIssue] | None = None

    path = f"{FOLDER_TEMP}/github_issues_{name}.json"

    # Read existing data
    existing_data_dict: TGitHubIssueExport = import_json(path)

    if not existing_data_dict:
        return result

    if str(GITHUB_MAX_PER_PAGE) in existing_data_dict:
        obj = existing_data_dict[str(GITHUB_MAX_PER_PAGE)]
        if str(page) in obj:
            arr = obj[str(page)]
            # if arr:
            result = [IGitHubIssue().model_validate(item) for item in arr]
            # print(f"{name}.{GITHUB_MAX_PER_PAGE}.{page} arr: {len(arr)}")

    return result


def export_json(
    data: any,
    filename: str | None = None,
    path: str | None = None,
    indent: int | None = None,
) -> str:

    if path:
        file_path = path
    else:
        # Create output directory if it doesn't exist
        Path(FOLDER_OUTPUT).mkdir(parents=True, exist_ok=True)
        file_path = f"{FOLDER_OUTPUT}/{filename}"

    # Convert data to JSON string if it isn't already
    if not isinstance(data, str):
        data = json.dumps(data, indent=indent, default=str)

    with open(file_path, "w", encoding="utf-8") as f:
        f.write(data)

    # print(f'🗂️  JSON_EXPORTED "{file_path}"')  # data:\n{data}

    return file_path


def export_json_issues(name: str, data: list[IGitHubIssue], page: int) -> str | None:

    path = f"{FOLDER_TEMP}/github_issues_{name}.json"

    # Read existing data or initialize an empty dict
    existing_data_dict: TGitHubIssueExport = import_json(path) or {}

    # Convert the list of IGitHubIssue to a list of dictionaries
    data_dict = [item.model_dump(by_alias=True) for item in data]

    # Update the dictionary for this max per page number
    if str(GITHUB_MAX_PER_PAGE) not in existing_data_dict:
        existing_data_dict[str(GITHUB_MAX_PER_PAGE)] = {}
    existing_data_dict[str(GITHUB_MAX_PER_PAGE)][str(page)] = data_dict

    output = export_json(data=existing_data_dict, path=path, indent=2)

    return output


def export_json_comments(
    name: str, issue_no: int, data: list[IGitHubComment]
) -> str | None:
    path = f"{FOLDER_TEMP}/github_issues_comments_{name}.json"
    # Read existing data or initialize an empty dict
    existing_data_dict: TGitHubIssueCommentExport = import_json(path) or {}

    # Convert the list of IGitHubComment to a list of dictionaries
    data_dict = [item.model_dump(by_alias=True) for item in data]

    # Update the dictionary for this issue number
    existing_data_dict[str(issue_no)] = data_dict

    # Write the entire dictionary back to the file
    output = export_json(data=existing_data_dict, path=path, indent=2)
    return output


def export_csv(
    data: List[Union[Dict[str, Any], BaseModel]],
    file_path: str,
    fieldnames: List[str] = None,
    delimiter: str = ",",
) -> bool:
    """
    Export data to a CSV file.

    Args:
        data: List of dictionaries or Pydantic model instances.
        file_path: Path to the output CSV file.
        fieldnames: Optional; list of field names. If not provided, uses keys from the first dictionary.
        delimiter: Optional; delimiter for the CSV, defaults to comma.

    Returns:
        True if successful, False otherwise.
    """
    try:
        # Convert Pydantic models to dictionaries
        if data and isinstance(data[0], BaseModel):
            data_dicts = [item.model_dump() for item in data]
        else:
            data_dicts = data

        # If fieldnames are not provided, use the keys from the first item
        if fieldnames is None and data_dicts:
            fieldnames = list(data_dicts[0].keys())

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(
                csvfile,
                fieldnames=fieldnames,
                delimiter=delimiter,
                quoting=csv.QUOTE_ALL,
            )
            writer.writeheader()
            for row in data_dicts:
                writer.writerow(row)

        print(f'🗂️  CSV_EXPORT Total "{len(data)}" into "{file_path}"')
        return True
    except Exception as e:
        print(f"❌ ERROR_CSV_EXPORT: {e}")
        return False


def export_csv_issues(name: str) -> str | None:

    input_path = f"{FOLDER_TEMP}/github_issues_{name}.json"
    output_path = f"{FOLDER_OUTPUT}/github_issues_{name}.csv"

    # Load the JSON data
    data = import_json(input_path)
    if data is None:
        return False

    # Flatten the comments: all values are lists, so we combine them into one list
    all_issues = []

    max_page_data = {}
    if str(GITHUB_MAX_PER_PAGE) in data:
        max_page_data = data[str(GITHUB_MAX_PER_PAGE)]

    for issues in max_page_data.values():
        all_issues.extend(issues)

    # Export to CSV
    isExport = export_csv(all_issues, output_path)

    if isExport:
        return output_path

    return None


def export_csv_comments(name: str) -> str | None:

    input_path = f"{FOLDER_TEMP}/github_issues_comments_{name}.json"
    output_path = f"{FOLDER_OUTPUT}/github_issues_comments_{name}.csv"

    # Load the JSON data
    data = import_json(input_path)
    if data is None:
        return False

    # Flatten the comments: all values are lists, so we combine them into one list
    all_comments = []
    for comments in data.values():
        all_comments.extend(comments)

    # Export to CSV
    isExport = export_csv(all_comments, output_path)

    if isExport:
        return output_path

    return None
